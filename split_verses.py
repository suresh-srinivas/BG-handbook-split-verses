
#!/usr/bin/env python3
"""
split_verses.py — Cut an audio file into fixed-length "verses" (or custom timestamps).

Features:
  - Start at an offset (e.g., 30s) and slice N segments of length L.
  - Optional CSV with timings.
  - Optional fades to avoid clicks.
  - Optional ZIP of outputs.
  - Alternative mode: provide a timestamps CSV to define exact cuts.

Requires: Python 3.9+, pydub, and ffmpeg installed & on PATH.
"""
import argparse
import csv
import os
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from typing import Dict, List, Optional, Tuple
from pydub import AudioSegment

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None


@dataclass
class Segment:
    label: str
    start_ms: int
    end_ms: int

    @property
    def duration_ms(self) -> int:
        return max(0, self.end_ms - self.start_ms)
import zipfile

def parse_time(s: str) -> int:
    """
    Parse time like "75" (seconds), "01:15", or "01:15.250" into milliseconds.
    """
    s = s.strip()
    if ':' not in s:
        # seconds (possibly float)
        sec = float(s)
        return int(round(sec * 1000))
    parts = s.split(':')
    if len(parts) == 2:
        mm, ss = parts
        return int(round((int(mm) * 60 + float(ss)) * 1000))
    elif len(parts) == 3:
        hh, mm, ss = parts
        total_sec = int(hh) * 3600 + int(mm) * 60 + float(ss)
        return int(round(total_sec * 1000))
    else:
        raise ValueError(f"Unrecognized time format: {s}")

def mmss(ms: int) -> str:
    s = int(round(ms/1000))
    m = s // 60
    sec = s % 60
    return f"{m:02d}:{sec:02d}"


def sanitize_filename(name: str, fallback: str) -> str:
    base = name.strip() if name else ""
    allowed = []
    for ch in base:
        if ch.isalnum() or ch in ("_", "-", "."):
            allowed.append(ch)
        elif ch.isspace():
            allowed.append("_")
    cleaned = "".join(allowed).strip("._")
    return cleaned or fallback


def normalize_header(cell: str) -> str:
    return "".join(ch for ch in cell.lower() if ch.isalnum())

def grid_cuts(start_ms: int, count: int, length_ms: int, total_ms: int) -> List[Tuple[int, int]]:
    cuts = []
    for i in range(count):
        st = start_ms + i * length_ms
        en = st + length_ms
        st = max(0, min(st, total_ms))
        en = max(0, min(en, total_ms))
        if en <= st:
            break
        cuts.append((st, en))
    return cuts


def parse_excel_time(value) -> int:
    """Parse the Excel "Beginning"/"Ending" cell into milliseconds."""

    def minutes_seconds_from_string(text: str) -> Tuple[int, int]:
        text = text.strip()
        if not text:
            raise ValueError("Empty time value")
        if ":" in text:
            return divmod(parse_time(text), 1000)
        if "." in text:
            minutes_str, seconds_str = text.split(".", 1)
            minutes = int(minutes_str) if minutes_str else 0
            seconds_digits = ''.join(ch for ch in seconds_str if ch.isdigit())
            if not seconds_digits:
                seconds = 0
            elif len(seconds_digits) == 1:
                seconds = int(seconds_digits) * 10
            else:
                seconds = int(seconds_digits[:2])
            seconds = max(0, min(seconds, 59))
            return minutes * 60 + seconds, 0
        # Plain number — treat as minutes (float allowed)
        minutes_float = float(text)
        total_seconds = int(round(minutes_float * 60))
        return total_seconds, 0

    if value is None:
        raise ValueError("Missing time value in Excel sheet")

    if isinstance(value, timedelta):
        total_seconds = value.total_seconds()
        return int(round(total_seconds * 1000))

    if isinstance(value, time):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
        return int(round(total_seconds * 1000))

    if isinstance(value, datetime):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
        return int(round(total_seconds * 1000))

    text_value = str(value).strip()
    if not text_value:
        raise ValueError("Empty time value in Excel sheet")

    seconds_or_ms, remainder = minutes_seconds_from_string(text_value)
    if remainder == 0:
        total_seconds = seconds_or_ms
    else:
        total_seconds = seconds_or_ms + remainder / 1000.0
    return int(round(total_seconds * 1000))


def clamp_segment(start_ms: int, end_ms: int, total_ms: int) -> Tuple[int, int]:
    st = max(0, min(start_ms, total_ms))
    en = max(0, min(end_ms, total_ms))
    return st, en


def load_timestamps_excel(path: str, total_ms: int) -> Dict[str, List[Segment]]:
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required to read Excel timestamp files. Install it with `pip install openpyxl`."
        )

    wb = load_workbook(path, data_only=True)
    chapters: Dict[str, List[Segment]] = {}

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        header_row: Optional[Tuple] = None
        for row in rows_iter:
            if row and any(cell is not None and str(cell).strip() for cell in row):
                header_row = row
                break
        if header_row is None:
            continue

        headers = [normalize_header(str(cell)) if cell is not None else "" for cell in header_row]
        idx_label = None
        idx_begin = None
        idx_end = None
        for idx, header in enumerate(headers):
            if idx_label is None and header in {"chaptersloka", "verse", "sloka", "name", "label"}:
                idx_label = idx
            if idx_begin is None and header in {"beginning", "begin", "start"}:
                idx_begin = idx
            if idx_end is None and header in {"ending", "end", "finish"}:
                idx_end = idx
        if idx_label is None or idx_begin is None or idx_end is None:
            raise ValueError(
                f"Sheet '{ws.title}' must contain 'Chapter Sloka', 'Beginning', and 'Ending' columns."
            )

        segments: List[Segment] = []
        for row in rows_iter:
            if not row:
                continue
            label_cell = row[idx_label] if idx_label < len(row) else None
            start_cell = row[idx_begin] if idx_begin < len(row) else None
            end_cell = row[idx_end] if idx_end < len(row) else None

            if label_cell is None or str(label_cell).strip() == "":
                continue
            if start_cell is None or end_cell is None:
                continue

            st = parse_excel_time(start_cell)
            en = parse_excel_time(end_cell)
            st, en = clamp_segment(st, en, total_ms)
            if en <= st:
                continue
            label = str(label_cell).strip()
            segments.append(Segment(label=label, start_ms=st, end_ms=en))

        if segments:
            segments.sort(key=lambda seg: seg.start_ms)
            chapters[ws.title] = segments

    if not chapters:
        raise ValueError("No usable data found in Excel file.")

    return chapters
def load_timestamps_csv(path: str, total_ms: int) -> List[Tuple[int,int]]:
    """
    CSV with either headers or not. Expect either:
      start,end
    or
      start,duration
    Time formats accept seconds or mm:ss(.ms)
    """
    cuts = []
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)
    # try to detect header
    def find_column(header_norm: List[str], keywords: Tuple[str, ...]) -> Optional[int]:
        for idx, norm in enumerate(header_norm):
            for kw in keywords:
                if norm == kw or norm.startswith(kw):
                    return idx
        return None

    header_raw = [c.strip() for c in rows[0]] if rows else []
    header_norm = [normalize_header(c) for c in header_raw]
    idx_start: Optional[int] = None
    idx_end: Optional[int] = None
    idx_dur: Optional[int] = None

    def row_looks_like_header(row: List[str]) -> bool:
        return any(any(ch.isalpha() for ch in cell) for cell in row)

    start_row = 0
    if header_raw and row_looks_like_header(header_raw):
        start_row = 1
        idx_start = find_column(header_norm, ("start", "begin"))
        idx_end = find_column(header_norm, ("end", "stop", "finish"))
        idx_dur = find_column(header_norm, ("duration", "length", "dur"))

    for r in rows[start_row:]:
        row = [cell.strip() for cell in r]
        if not row or all(not cell for cell in row):
            continue
        use_end = False
        start_s: Optional[str] = None
        second: Optional[str] = None

        if idx_start is not None and idx_start < len(row) and row[idx_start]:
            start_s = row[idx_start]
            if idx_end is not None and idx_end < len(row) and row[idx_end]:
                second = row[idx_end]
                use_end = True
            elif idx_dur is not None and idx_dur < len(row) and row[idx_dur]:
                second = row[idx_dur]
        if start_s is None or second is None:
            # fall back to the first two populated columns (legacy behaviour)
            populated = [cell for cell in row if cell]
            if len(populated) < 2:
                raise ValueError("Timestamps CSV needs at least 2 columns (start,end or start,duration).")
            start_s, second = populated[0], populated[1]
            use_end = False

        st = parse_time(start_s)
        if use_end:
            en = parse_time(second)
        else:
            dur = parse_time(second)
            en = st + dur

        # clamp
        st = max(0, min(st, total_ms))
        en = max(0, min(en, total_ms))
        if en > st:
            cuts.append((st, en))
    return cuts

def main():
    ap = argparse.ArgumentParser(description="Split audio into verses by grid or timestamps.")
    ap.add_argument("-i","--input", required=True, help="Input audio file (mp3/wav/etc.)")
    ap.add_argument("-o","--output", default="verses_out", help="Output directory")
    ap.add_argument("--start", default="0", help='Start offset (seconds or mm:ss), default 0')
    ap.add_argument("--count", type=int, default=20, help="Number of verses to cut (grid mode)")
    ap.add_argument("--length", default="15", help="Verse length (seconds or mm:ss) in grid mode")
    ap.add_argument("--timestamps", help="CSV of custom cuts: start,end or start,duration (overrides grid)")
    ap.add_argument(
        "--timestamps-excel",
        dest="timestamps_excel",
        help="Excel workbook with one sheet per chapter (Chapter Sloka / Beginning / Ending)",
    )
    ap.add_argument("--prefix", default="Verse_", help="Filename prefix, default Verse_")
    ap.add_argument("--bitrate", default="192k", help="Output bitrate for mp3, default 192k")
    ap.add_argument("--fade_in", type=int, default=5, help="Fade in ms, default 5")
    ap.add_argument("--fade_out", type=int, default=10, help="Fade out ms, default 10")
    ap.add_argument("--zip", dest="make_zip", action="store_true", help="Also produce a ZIP of outputs")
    ap.add_argument("--csv", dest="csv_out", help="Write a timings CSV to this path")
    args = ap.parse_args()

    os.makedirs(args.output, exist_ok=True)

    audio = AudioSegment.from_file(args.input)
    total_ms = len(audio)

    if args.timestamps and args.timestamps_excel:
        raise SystemExit("Please provide only one of --timestamps or --timestamps-excel, not both.")

    export_paths: List[str] = []
    csv_rows: List[Tuple] = []
    csv_header: Optional[List[str]] = None

    if args.timestamps_excel:
        chapters = load_timestamps_excel(args.timestamps_excel, total_ms)
        used_dirnames: Dict[str, int] = {}

        for sheet_name, segments in chapters.items():
            base_dirname = sanitize_filename(sheet_name, "chapter")
            count = used_dirnames.get(base_dirname, 0)
            if count:
                dirname = f"{base_dirname}_{count+1}"
            else:
                dirname = base_dirname
            used_dirnames[base_dirname] = count + 1

            sheet_dir = os.path.join(args.output, dirname)
            os.makedirs(sheet_dir, exist_ok=True)

            for idx, seg in enumerate(segments, start=1):
                fallback_name = f"{args.prefix}{idx:02d}"
                base_fname = sanitize_filename(seg.label, fallback_name)
                fname = f"{base_fname}.mp3"
                fpath = os.path.join(sheet_dir, fname)
                suffix = 2
                while os.path.exists(fpath):
                    fname = f"{base_fname}_{suffix}.mp3"
                    fpath = os.path.join(sheet_dir, fname)
                    suffix += 1

                seg_audio = audio[seg.start_ms:seg.end_ms].fade_in(args.fade_in).fade_out(args.fade_out)
                seg_audio.export(fpath, format="mp3", bitrate=args.bitrate)
                export_paths.append(fpath)

                csv_rows.append(
                    (
                        sheet_name,
                        seg.label,
                        mmss(seg.start_ms),
                        mmss(seg.end_ms),
                        round(seg.duration_ms / 1000, 3),
                        os.path.relpath(fpath, args.output),
                    )
                )

        csv_header = ["Chapter", "Verse", "Start", "End", "Duration(s)", "File"]

    elif args.timestamps:
        cuts = load_timestamps_csv(args.timestamps, total_ms)
        if not cuts:
            raise SystemExit("No valid cuts parsed from timestamps CSV.")
        for idx, (st, en) in enumerate(cuts, start=1):
            seg = audio[st:en].fade_in(args.fade_in).fade_out(args.fade_out)
            fname = f"{args.prefix}{idx:02d}.mp3"
            fpath = os.path.join(args.output, fname)
            seg.export(fpath, format="mp3", bitrate=args.bitrate)
            export_paths.append(fpath)
            csv_rows.append((idx, mmss(st), mmss(en), round(len(seg)/1000, 3), fname))

        csv_header = ["Verse", "Start", "End", "Duration(s)", "File"]

    else:
        start_ms = parse_time(args.start)
        length_ms = parse_time(args.length)
        cuts = grid_cuts(start_ms, args.count, length_ms, total_ms)
        if not cuts:
            raise SystemExit("No valid cuts produced by grid. Check --start/--count/--length.")
        for idx, (st, en) in enumerate(cuts, start=1):
            seg = audio[st:en].fade_in(args.fade_in).fade_out(args.fade_out)
            fname = f"{args.prefix}{idx:02d}.mp3"
            fpath = os.path.join(args.output, fname)
            seg.export(fpath, format="mp3", bitrate=args.bitrate)
            export_paths.append(fpath)
            csv_rows.append((idx, mmss(st), mmss(en), round(len(seg)/1000, 3), fname))

        csv_header = ["Verse", "Start", "End", "Duration(s)", "File"]

    if args.csv_out:
        with open(args.csv_out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if csv_header:
                w.writerow(csv_header)
            for r in csv_rows:
                w.writerow(r)

    if args.make_zip:
        zip_path = os.path.join(args.output, "verses.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in export_paths:
                arcname = os.path.relpath(p, args.output)
                zf.write(p, arcname=arcname)

    print(f"Done. Wrote {len(export_paths)} files to: {args.output}")
    if args.make_zip:
        print(f"ZIP: {os.path.join(args.output, 'verses.zip')}")
    if args.csv_out:
        print(f"CSV: {args.csv_out}")

if __name__ == "__main__":
    main()
